# -*- coding: utf-8 -*-


from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.styles import Border,Side


def get_py_dtl_sum_param_desc(lang, py_sum=False):
    """获取薪资明细表/汇总表参数描述"""

    fixed_parm_dic = {}
    if lang == 'zh_CN':
        if py_sum:
            fixed_parm_dic['title_tl'] = '薪资汇总表'
        else:
            fixed_parm_dic['title_tl'] = '薪资明细表'
        fixed_parm_dic['yes_tl'] = '是'
        fixed_parm_dic['no_tl'] = '否'
        fixed_parm_dic['cal_tl'] = '日历'
        fixed_parm_dic['pay_group_tl'] = '薪资组'
        fixed_parm_dic['begin_date_tl'] = '开始日期'
        fixed_parm_dic['end_date_tl'] = '结束日期'
        fixed_parm_dic['dept_tl'] = '部门'
        fixed_parm_dic['has_child_dept_lbl'] = '是否包含子部门'
        fixed_parm_dic['emp_tl'] = '人员'
        fixed_parm_dic['rcd_tl'] = '任职记录号'
        fixed_parm_dic['col_trac_detail_lbl'] = '分列显示追溯明细'
        fixed_parm_dic['emp_id_tl'] = '人员编号'
        fixed_parm_dic['name_tl'] = '姓名'
        fixed_parm_dic['dept_name_tl'] = '部门名称'
        fixed_parm_dic['dept_full_name'] = '部门全称'
        fixed_parm_dic['pos_name_tl'] = '职位名称'
        fixed_parm_dic['tree_lvl_lbl'] = '按树层级汇总'
    else:
        if py_sum:
            fixed_parm_dic['title_tl'] = 'Payroll Summary'
        else:
            fixed_parm_dic['title_tl'] = 'Payroll Detail'
        fixed_parm_dic['yes_tl'] = 'Yes'
        fixed_parm_dic['no_tl'] = 'No'
        fixed_parm_dic['cal_tl'] = 'Calendar'
        fixed_parm_dic['pay_group_tl'] = 'Payroll Group'
        fixed_parm_dic['begin_date_tl'] = 'Begin Date'
        fixed_parm_dic['end_date_tl'] = 'End Date'
        fixed_parm_dic['dept_tl'] = 'Department'
        fixed_parm_dic['has_child_dept_lbl'] = 'Include Sub-departments'
        fixed_parm_dic['emp_tl'] = 'Employee ID'
        fixed_parm_dic['rcd_tl'] = 'Employee Record No.'
        fixed_parm_dic['col_trac_detail_lbl'] = 'Column-display retro details'
        fixed_parm_dic['emp_id_tl'] = 'Employee ID'
        fixed_parm_dic['name_tl'] = 'Name'
        fixed_parm_dic['dept_name_tl'] = 'Department Name'
        fixed_parm_dic['dept_full_name'] = 'Department Full Name'
        fixed_parm_dic['pos_name_tl'] = 'Position Name'
        fixed_parm_dic['tree_lvl_lbl'] = 'Group by tree level'
    return fixed_parm_dic


def get_num_chn(num):
    num_chn_dic = {1: '一', 2: '二', 3: '三', 4: '四', 5: '五', 6: '六', 7: '七', 8: '八', 9: '九',
                   10: '十', 11: '十一', 12: '十二', 13: '十三', 14: '十四', 15: '十五', 16: '十六', 17: '十七',
                   18: '十八', 19: '十九', 20: '二十', 21: '二十一', 22: '二十二', 23: '二十三', 24: '二十四', 25: '二十五', 26: '二十六',
                   27: '二十七', 28: '二十八', 29: '二十九', 30: '三十'}
    if num in num_chn_dic:
        return num_chn_dic[num]
    else:
        return ''


def get_dept_lvl_desc_dic(lang):
    dept_lvl_desc_dic = {}
    # 预置30个部门层级
    for i in range(1, 31):
        if lang == 'zh_CN':
            dept_lvl_desc_dic[i] = get_num_chn(i) + '级部门'
        else:
            dept_lvl_desc_dic[i] = 'Level ' + str(i) + ' Department'
    return dept_lvl_desc_dic


def create_py_detail_excl(field_tab, result_list, run_param):
    """
    Desc: 生成薪资明细表Excel
    Author: 陶雨
    Date: 2019/01/17
    """

    fixed_parm_dic = get_py_dtl_sum_param_desc(run_param['lang'])
    title_tl = fixed_parm_dic['title_tl']
    yes_tl = fixed_parm_dic['yes_tl']
    no_tl = fixed_parm_dic['no_tl']
    cal_tl = fixed_parm_dic['cal_tl']
    pay_group_tl = fixed_parm_dic['pay_group_tl']
    begin_date_tl = fixed_parm_dic['begin_date_tl']
    end_date_tl = fixed_parm_dic['end_date_tl']
    dept_tl = fixed_parm_dic['dept_tl']
    has_child_dept_lbl = fixed_parm_dic['has_child_dept_lbl']
    emp_tl = fixed_parm_dic['emp_tl']
    rcd_tl = fixed_parm_dic['rcd_tl']
    col_trac_detail_lbl = fixed_parm_dic['col_trac_detail_lbl']
    emp_id_tl = fixed_parm_dic['emp_id_tl']
    name_tl = fixed_parm_dic['name_tl']
    dept_name_tl = fixed_parm_dic['dept_name_tl']
    dept_full_name = fixed_parm_dic['dept_full_name']
    pos_name_tl = fixed_parm_dic['pos_name_tl']

    wb = Workbook()
    sheet = wb.active
    sheet.title = title_tl
    rows = []
    if run_param['has_child_dept'] == 'Y':
        has_child_dept = yes_tl
    else:
        has_child_dept = no_tl
    if run_param['col_trac_detail'] == 'Y':
        col_trac_detail = yes_tl
    else:
        col_trac_detail = no_tl
    excel_file_path = run_param['excel_file_path']
    row_para1 = [cal_tl, run_param['py_cal_id']]
    row_para2 = [pay_group_tl, run_param['py_group_id']]
    row_para3 = [begin_date_tl, run_param['start_date']]
    row_para4 = [end_date_tl, run_param['end_date']]
    row_para5 = [dept_tl, run_param['dept_cd']]
    row_para6 = [has_child_dept_lbl, has_child_dept]
    row_para7 = [emp_tl, run_param['emp_id']]
    row_para8 = [rcd_tl, run_param['emp_rcd']]
    row_para9 = [col_trac_detail_lbl, col_trac_detail]
    row_blank = []
    row_head = [emp_id_tl, name_tl, rcd_tl, begin_date_tl, end_date_tl, pay_group_tl, cal_tl]
    rows.append(row_para1)
    rows.append(row_para2)
    rows.append(row_para3)
    rows.append(row_para4)
    rows.append(row_para5)
    rows.append(row_para6)
    rows.append(row_para7)
    rows.append(row_para8)
    rows.append(row_para9)
    rows.append(row_blank)

    if field_tab is not None:
        for tab_desc in field_tab:
            # 如果当前标签是部门或者职位
            if '##' in tab_desc:
                tab_desc_lst = tab_desc.split('##')
                tab_desc = tab_desc_lst[0]
                field_name = tab_desc_lst[1]
                if field_name == 'HHR_TXT001':
                    row_head.append(tab_desc)
                    row_head.append(dept_name_tl)
                    row_head.append(dept_full_name)
                elif field_name == 'HHR_TXT002':
                    row_head.append(tab_desc)
                    row_head.append(pos_name_tl)
            else:
                row_head.append(tab_desc)

    rows.append(row_head)
    if result_list is not None:
        for row_data in result_list:
            rows.append(row_data)
    # 添加
    for addrow in rows:
        sheet.append(addrow)

    # 参数字体加粗
    for i in range(1, 10):
        sheet['A' + str(i)].font = Font(bold=True)

    # 遍历每一列
    for i in range(1, sheet.max_column + 1):
        column_name = get_column_letter(i)
        # 设置每列的宽度
        sheet.column_dimensions[column_name].width = 25
        # 表头加粗
        sheet[column_name + '11'].font = Font(bold=True)

    # 去掉网格线
    # sheet.sheet_view.showGridLines = False
    # 冻结窗口
    sheet.freeze_panes = sheet['A12']
    wb.save(excel_file_path)


def create_py_sum_excl(max_h_dept_cnt, field_tab, result_list, header_inx_lst, run_param):
    """
    Desc: 生成薪资汇总Excel
    Author: David
    Date: 2019/03/05
    """

    lang = run_param['lang']
    fixed_parm_dic = get_py_dtl_sum_param_desc(lang, py_sum=True)
    title_tl = fixed_parm_dic['title_tl']
    yes_tl = fixed_parm_dic['yes_tl']
    no_tl = fixed_parm_dic['no_tl']
    cal_tl = fixed_parm_dic['cal_tl']
    pay_group_tl = fixed_parm_dic['pay_group_tl']
    begin_date_tl = fixed_parm_dic['begin_date_tl']
    end_date_tl = fixed_parm_dic['end_date_tl']
    dept_tl = fixed_parm_dic['dept_tl']
    has_child_dept_lbl = fixed_parm_dic['has_child_dept_lbl']
    tree_lvl_lbl = fixed_parm_dic['tree_lvl_lbl']
    dept_name_tl = fixed_parm_dic['dept_name_tl']
    dept_full_name = fixed_parm_dic['dept_full_name']
    pos_name_tl = fixed_parm_dic['pos_name_tl']

    wb = Workbook()
    sheet = wb.active
    sheet.title = title_tl
    rows = []
    if run_param['has_child_dept'] == 'Y':
        has_child_dept = yes_tl
    else:
        has_child_dept = no_tl

    excel_file_path = run_param['excel_file_path']
    row_para1 = [cal_tl, run_param['py_cal_id']]
    row_para2 = [pay_group_tl, run_param['py_group_id']]
    row_para3 = [begin_date_tl, run_param['start_date']]
    row_para4 = [end_date_tl, run_param['end_date']]
    row_para5 = [dept_tl, run_param['dept_cd']]
    row_para6 = [has_child_dept_lbl, has_child_dept]
    row_para7 = [tree_lvl_lbl, run_param['tree_lvl']]
    row_blank = []

    row_head = [begin_date_tl, end_date_tl]

    # 显示各部门各层级描述
    dept_lvl_desc_dic = get_dept_lvl_desc_dic(lang)
    for j in range(1, max_h_dept_cnt + 1):
        if j in dept_lvl_desc_dic:
            row_head.append(dept_lvl_desc_dic[j])

    rows.append(row_para1)
    rows.append(row_para2)
    rows.append(row_para3)
    rows.append(row_para4)
    rows.append(row_para5)
    rows.append(row_para6)
    rows.append(row_para7)
    rows.append(row_blank)

    if field_tab is not None:
        for tab_desc in field_tab:
            # 如果当前标签是部门或者职位
            if '##' in tab_desc:
                tab_desc_lst = tab_desc.split('##')
                tab_desc = tab_desc_lst[0]
                field_name = tab_desc_lst[1]
                if field_name == 'HHR_TXT001':
                    row_head.append(tab_desc)
                    row_head.append(dept_name_tl)
                    row_head.append(dept_full_name)
                elif field_name == 'HHR_TXT002':
                    row_head.append(tab_desc)
                    row_head.append(pos_name_tl)
            else:
                row_head.append(tab_desc)

    # 如果按树层级进行了汇总，还需要过滤掉不需要的列
    if len(header_inx_lst) > 0:
        new_row_head = []
        for k in range(len(row_head)):
            if k in header_inx_lst:
                new_row_head.append(row_head[k])
        row_head = new_row_head

    rows.append(row_head)
    if result_list is not None:
        for row_data in result_list:
            rows.append(row_data)
    # 添加
    for addrow in rows:
        sheet.append(addrow)

    # 参数字体加粗
    for i in range(1, 8):
        sheet['A' + str(i)].font = Font(bold=True)

    # 遍历每一列
    for i in range(1, sheet.max_column + 1):
        column_name = get_column_letter(i)
        # 设置每列的宽度
        sheet.column_dimensions[column_name].width = 25
        # 表头加粗
        sheet[column_name + '9'].font = Font(bold=True)

    # 去掉网格线
    # sheet.sheet_view.showGridLines = False
    # 冻结窗口
    sheet.freeze_panes = sheet['A10']
    wb.save(excel_file_path)


def create_pers_trans_excl(result_list, run_param):
    """
    Desc: 生成人员异动情况Excel
    Author: 陶雨
    Date: 2019/04/23
    """
    excel_file_path = run_param['excel_file_path']
    wb = Workbook()
    sheet = wb.active
    sheet.title = "人员异动情况表"
    rows = []
    row0 = ['集团', '公司', '一级部门', '二级部门', '三级部门', '四级部门', '总人数', '晋升人数', '降职人数', '入职人数', '离职人数', '调入人数', '调出人数', '增加人数', '减少人数']
    rows.append(row0)
    for row_data in result_list:
        rows.append(row_data)
    for addrow in rows:
        sheet.append(addrow)

    # 遍历每一列
    for i in range(1, sheet.max_column + 1):
        column_name = get_column_letter(i)
        # 设置每列的宽度
        sheet.column_dimensions[column_name].width = 25
        # 表头底色
        sheet[column_name + '1'].fill = PatternFill(start_color="C1C0BF", end_color="C1C0BF", fill_type="solid")

    wb.save(excel_file_path)
